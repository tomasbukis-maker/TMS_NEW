"""Konvertuoja senos sistemos "infotrans_uzsakymai.xlsx" į normalizuotą CSV.

Vykdymas:
    python tools/importers/convert_infotrans_orders.py \
        --input frontend/public/infotrans_uzsakymai.xlsx \
        --output frontend/public/infotrans_uzsakymai_converted.csv

Naudoja openpyxl, nes pandas šiame aplinkoje neįdiegtas.
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import load_workbook


DATE_PATTERN = re.compile(r"(\d{4})[.\-/](\d{2})[.\-/](\d{2})")
CARGO_EP_PATTERN = re.compile(r"(\d+[.,]?\d*)\s*EP", re.IGNORECASE)
CARGO_LDM_PATTERN = re.compile(r"(\d+[.,]?\d*)\s*LDM", re.IGNORECASE)
CARGO_WEIGHT_PATTERN = re.compile(r"(\d+[.,]?\d*)\s*KG", re.IGNORECASE)
EXPEDITION_PATTERN = re.compile(r"^([A-Za-z]+\d+)\s*(.*)$")

COUNTRY_TOKEN = re.compile(r"^[A-Z]{2}$")
POSTAL_TOKEN = re.compile(r"^[A-Z]{0,2}-?\d{3,6}$")

INVOICE_DATE_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}|\d{4}\.\d{2}\.\d{2}|\d{2}\.\d{2}\.\d{2}")

OUTPUT_HEADERS = [
    "order_number",
    "client_name",
    "client_name_clean",
    "loading_date",
    "unloading_date",
    "route_from_country",
    "route_from_postal_code",
    "route_from_city",
    "route_to_country",
    "route_to_postal_code",
    "route_to_city",
    "cargo_ep",
    "cargo_ldm",
    "cargo_weight_kg",
    "cargo_description",
    "client_amount_eur",
    "carrier_amount_eur",
    "profit_eur",
    "sales_invoices",
    "sales_invoice_numbers",
    "sales_invoice_terms",
    "sales_invoice_dates",
    "purchase_invoices",
    "purchase_invoice_numbers",
    "purchase_invoice_terms",
    "purchase_invoice_dates",
    "expedition_number",
    "carrier_details",
    "carrier_details_clean",
    "notes",
]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Konvertuoti Infotrans užsakymų Excel failą.")
    parser.add_argument("--input", required=True, type=Path, help="Kelias iki originalaus XLSX failo")
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Kelias iki sugeneruojamo CSV (bus perrašytas)",
    )
    return parser.parse_args()


def find_header(rows: Iterable[Tuple]) -> Tuple[int, List[str]]:
    """Randa pirmą ne tuščią eilutę (antraštę) ir grąžina jos indeksą bei tekstines reikšmes."""
    for idx, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() for cell in row):
            header = ["" if cell is None else str(cell).strip() for cell in row]
            return idx, header
    raise ValueError("Nepavyko rasti antraštės eilutės")


def normalise_header(header: List[str]) -> List[str]:
    """Patvarko antraštės pavadinimus, pašalina perteklinius tarpus."""
    return [h.strip() if h else "" for h in header]


def find_column(header: List[str], name: str) -> int:
    try:
        return header.index(name)
    except ValueError as exc:
        raise ValueError(f'Nepavyko rasti stulpelio "{name}"') from exc


def parse_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    match = DATE_PATTERN.search(text)
    if match:
        year, month, day = match.groups()
        return f"{year}-{month}-{day}"
    return ""


def parse_amount(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float, Decimal)):
        return f"{Decimal(str(value)).quantize(Decimal('0.01'))}"
    text = str(value)
    text = text.replace("EUR", "").replace("€", "").replace(" ", "").replace(" ", "")
    text = text.replace(",", ".")
    text = text.strip()
    if not text:
        return ""
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return ""
    return f"{amount.quantize(Decimal('0.01'))}"


def parse_cargo(value) -> Tuple[str, str, str, str]:
    if not value:
        return "", "", "", ""
    text = str(value).strip()
    ep = _extract_number(CARGO_EP_PATTERN, text)
    ldm = _extract_number(CARGO_LDM_PATTERN, text)
    weight = _extract_number(CARGO_WEIGHT_PATTERN, text)
    return ep, ldm, weight, text


def _extract_number(pattern: re.Pattern, text: str) -> str:
    match = pattern.search(text)
    if not match:
        return ""
    value = match.group(1).replace(",", ".")
    try:
        number = Decimal(value)
    except InvalidOperation:
        return ""
    return _decimal_to_plain(number)


def _decimal_to_plain(number: Decimal) -> str:
    normalized = number.normalize()
    text = format(normalized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def split_expedition(value) -> Tuple[str, str]:
    if not value:
        return "", ""
    text = str(value).strip()
    match = EXPEDITION_PATTERN.match(text)
    if match:
        return match.group(1), match.group(2).strip()
    return "", text


def split_route_components(value) -> Tuple[str, str, str]:
    if not value:
        return "", "", ""
    text = str(value).strip().strip("-").strip()
    if not text:
        return "", "", ""
    tokens = [token for token in re.split(r"[\s,]+", text) if token]
    country = ""
    postal = ""
    remaining: List[str] = []
    for token in tokens:
        upper = token.upper()
        if not country and COUNTRY_TOKEN.fullmatch(upper):
            country = upper
            continue
        if country and COUNTRY_TOKEN.fullmatch(upper) and upper == country:
            # pasikartojantis šalies kodas
            continue
        if not postal and POSTAL_TOKEN.fullmatch(upper):
            postal = upper
            continue
        remaining.append(token)
    city = " ".join(remaining).strip("- ")
    return country, postal, city


def _normalise_invoice_entries(value) -> Tuple[str, str, str]:
    entries = parse_invoice_entries(value)
    if not entries:
        return "", "", ""
    numbers = []
    terms = []
    dates = []
    for entry in entries:
        number, term, date = entry
        numbers.append(number)
        terms.append(term)
        dates.append(date)
    return ";".join(numbers), ";".join(filter(None, terms)), ";".join(filter(None, dates))


def parse_invoice_entries(value) -> List[Tuple[str, str, str]]:
    if not value:
        return []
    text = str(value).strip()
    if not text:
        return []
    text = re.sub(r"\s+", " ", text)
    entries: List[Tuple[str, str, str]] = []
    last_end = 0
    for match in INVOICE_DATE_PATTERN.finditer(text):
        segment = text[last_end : match.start()].strip()
        number, term = _split_invoice_segment(segment)
        date_clean = normalise_invoice_date(match.group())
        if number or term or date_clean:
            entries.append((number, term, date_clean))
        last_end = match.end()

    tail = text[last_end:].strip()
    if tail:
        number, term = _split_invoice_segment(tail)
        if number or term:
            entries.append((number, term, ""))

    if not entries:
        return [(text, "", "")]
    return entries


def _split_invoice_segment(segment: str) -> Tuple[str, str]:
    if not segment:
        return "", ""
    tokens = segment.split()
    if not tokens:
        return "", ""
    if len(tokens) >= 2 and tokens[-2].isdigit() and tokens[-1].lower().startswith("d"):
        term = tokens[-2].lstrip("0") or "0"
        number = " ".join(tokens[:-2])
        return number.strip(), term
    if tokens[-1].lower().startswith("d") and len(tokens) >= 2 and tokens[-2].isdigit():
        term = tokens[-2].lstrip("0") or "0"
        number = " ".join(tokens[:-2])
        return number.strip(), term
    return segment.strip(), ""


def normalise_invoice_date(value: str) -> str:
    if not value:
        return ""
    value = value.strip()
    if not value:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return value
    if re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", value):
        year, month, day = value.split(".")
        return f"{year}-{month}-{day}"
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", value):
        year, month, day = value.split(".")
        year_int = int(year)
        year_full = 2000 + year_int if year_int < 70 else 1900 + year_int
        return f"{year_full:04d}-{month}-{day}"
    return ""


def clean_name(value: str) -> str:
    if not value:
        return ""
    text = str(value)
    text = text.replace("..", "").replace("…", "")
    text = text.strip()
    return re.sub(r"\s+", " ", text)


def clean_row(row: Tuple, header: List[str]) -> dict:
    column = {header[idx]: row[idx] if idx < len(row) else None for idx in range(len(header))}

    order_number = (column.get("Uzsak. Nr.") or "").strip()
    if not order_number:
        return {}

    cargo_ep, cargo_ldm, cargo_weight, cargo_description = parse_cargo(column.get("EP,Ldm,Svoris"))
    expedition_number, carrier_details = split_expedition(column.get("Ekspedicijos nr, Vezejas, papildomos islaidos"))

    route_from_country, route_from_postal, route_from_city = split_route_components(column.get("Marsrutas  is"))
    route_to_country, route_to_postal, route_to_city = split_route_components(column.get("Maršrutas į"))

    sales_numbers, sales_terms, sales_dates = _normalise_invoice_entries(column.get("Israsytos saskaitos"))
    purchase_numbers, purchase_terms, purchase_dates = _normalise_invoice_entries(column.get("Gautos saskaitos"))

    return {
        "order_number": order_number,
        "client_name": (column.get("Uzsakovas") or "").strip(),
        "client_name_clean": clean_name(column.get("Uzsakovas")),
        "loading_date": parse_date(column.get("Pakrovimo data")),
        "unloading_date": parse_date(column.get("Iskrovima data")),
        "route_from_country": route_from_country,
        "route_from_postal_code": route_from_postal,
        "route_from_city": route_from_city,
        "route_to_country": route_to_country,
        "route_to_postal_code": route_to_postal,
        "route_to_city": route_to_city,
        "cargo_ep": cargo_ep,
        "cargo_ldm": cargo_ldm,
        "cargo_weight_kg": cargo_weight,
        "cargo_description": cargo_description,
        "client_amount_eur": parse_amount(column.get("Suma uzsakovui")),
        "carrier_amount_eur": parse_amount(column.get("Suma vezejui")),
        "profit_eur": parse_amount(column.get("Pelnas EUR")),
        "sales_invoices": (column.get("Israsytos saskaitos") or "").strip(),
        "sales_invoice_numbers": sales_numbers,
        "sales_invoice_terms": sales_terms,
        "sales_invoice_dates": sales_dates,
        "purchase_invoices": (column.get("Gautos saskaitos") or "").strip(),
        "purchase_invoice_numbers": purchase_numbers,
        "purchase_invoice_terms": purchase_terms,
        "purchase_invoice_dates": purchase_dates,
        "expedition_number": expedition_number,
        "carrier_details": carrier_details,
        "carrier_details_clean": clean_name(carrier_details),
        "notes": (column.get("Pastabos") or "").strip(),
    }


def convert(input_path: Path, output_path: Path) -> Tuple[int, int]:
    wb = load_workbook(filename=input_path, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel faile nėra lapų")

    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx, header = find_header(all_rows)
    header = normalise_header(header)

    data_rows = []
    skipped = 0
    for row in all_rows[header_idx + 1 :]:
        cleaned = clean_row(row, header)
        if not cleaned:
            skipped += 1
            continue
        data_rows.append(cleaned)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        for row in data_rows:
            writer.writerow(row)

    return len(data_rows), skipped


def main() -> None:
    args = parse_arguments()
    written, skipped = convert(args.input, args.output)
    print(f"Iš viso įrašyta: {written} | praleista tuščių/ar netinkamų eilučių: {skipped}")


if __name__ == "__main__":
    main()

