import csv
import os
from django.core.management.base import BaseCommand
from apps.partners.models import Partner, Contact


class Command(BaseCommand):
    help = 'Importuoja klientus iš CSV arba XLSX failo'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Kelias į CSV arba XLSX failą')
        parser.add_argument(
            '--is-client',
            action='store_true',
            help='Pažymėti visus partnerius kaip klientus',
        )
        parser.add_argument(
            '--is-supplier',
            action='store_true',
            help='Pažymėti visus partnerius kaip tiekėjus',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        is_client = options.get('is_client', True)  # Default - klientai
        is_supplier = options.get('is_supplier', False)
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Failas nerastas: {file_path}'))
            return
        
        imported_count = 0
        error_count = 0
        
        # Nustatome failo tipą
        is_xlsx = file_path.lower().endswith('.xlsx')
        
        try:
            # Apdorojame XLSX failą
            if is_xlsx:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.stdout.write(self.style.ERROR(
                        'openpyxl biblioteka nerasta. Įdiekite: pip install openpyxl'
                    ))
                    return
                
                workbook = load_workbook(file_path)
                sheet = workbook.active
                
                # Gauname header eilutę
                headers = []
                for cell in sheet[1]:
                    headers.append(cell.value if cell.value else '')
                
                # Konvertuojame į dict formatą
                rows_data = []
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for idx, cell in enumerate(row):
                        header_name = headers[idx] if idx < len(headers) else ''
                        row_dict[header_name] = str(cell.value) if cell.value else ''
                    rows_data.append(row_dict)
                
                rows = rows_data
            else:
                # Apdorojame CSV failą
                with open(file_path, 'r', encoding='utf-8') as file:
                    # Bandoma nustatyti delimiter (gali būti ; arba ,)
                    sample = file.read(1024)
                    file.seek(0)
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                    
                    reader = csv.DictReader(file, delimiter=delimiter)
                    rows = list(reader)
            
            # Helper funkcija duomenims parsinti
            def safe_get(row, key, default=''):
                value = row.get(key, default)
                return str(value).strip() if value is not None else default
            
            # Apdorojame eilutes
            for row_num, row in enumerate(rows, start=2):  # prasideda nuo 2 (1 yra header)
                    try:
                        # Parsiname duomenis
                        firmos_pavadinimas = safe_get(row, 'Firmos pavadinimas')
                        if not firmos_pavadinimas:
                            self.stdout.write(self.style.WARNING(f'Eilutė {row_num}: Praleista - nėra firmos pavadinimo'))
                            continue
                        
                        imones_kodas = safe_get(row, 'Imones kodas')
                        if not imones_kodas:
                            self.stdout.write(self.style.WARNING(f'Eilutė {row_num}: Praleista - nėra įmonės kodo'))
                            continue
                        
                        # Tikriname ar partneris jau egzistuoja
                        if Partner.objects.filter(code=imones_kodas).exists():
                            self.stdout.write(self.style.WARNING(
                                f'Eilutė {row_num}: Partneris su kodu {imones_kodas} jau egzistuoja - praleista'
                            ))
                            continue
                        
                        # Formuojame adresą
                        salis = safe_get(row, 'Salis')
                        miestas = safe_get(row, 'Miestas')
                        address_parts = [miestas, salis]
                        address = ', '.join([p for p in address_parts if p])
                        
                        # Sukuriame kontaktinį asmenį, jei yra duomenų
                        contact_person = None
                        kontaktiniai_asmenys = safe_get(row, 'Kontaktiniai asmenys')
                        el_pastas = safe_get(row, 'El. pastas')
                        telefonas = safe_get(row, 'Telefonas')
                        
                        if kontaktiniai_asmenys or el_pastas or telefonas:
                            # Bandoma parse'inti kontaktinius asmenis (pvz., "Vardas Pavardė" arba "Vardas; Pavardė")
                            name_parts = kontaktiniai_asmenys.replace(';', ',').split(',') if kontaktiniai_asmenys else []
                            first_name = name_parts[0].strip() if len(name_parts) > 0 else ''
                            last_name = name_parts[1].strip() if len(name_parts) > 1 else kontaktiniai_asmenys.strip() if not first_name else ''
                            
                            # Jei nėra vardo, naudojame firmos pavadinimą arba "Nenurodytas"
                            if not first_name and not last_name:
                                first_name = 'Nenurodytas'
                            
                            contact_person = Contact.objects.create(
                                first_name=first_name,
                                last_name=last_name,
                                email=el_pastas if el_pastas else '',
                                phone=telefonas if telefonas else '',
                            )
                        
                        # Sukuriame partnerį
                        partner = Partner.objects.create(
                            name=firmos_pavadinimas,
                            code=imones_kodas,
                            vat_code=safe_get(row, 'PVM kodas'),
                            address=address,
                            contact_person=contact_person,
                            is_client=is_client,
                            is_supplier=is_supplier,
                            status='active',
                            payment_term_days=30,  # Default
                        )
                        
                        imported_count += 1
                        self.stdout.write(self.style.SUCCESS(
                            f'Eilutė {row_num}: Importuotas - {firmos_pavadinimas} ({imones_kodas})'
                        ))
                        
                    except Exception as e:
                        error_count += 1
                        self.stdout.write(self.style.ERROR(
                            f'Eilutė {row_num}: Klaida - {str(e)}'
                        ))
                        continue
                
            self.stdout.write(self.style.SUCCESS(
                f'\n✓ Importuota: {imported_count} partnerių'
            ))
            if error_count > 0:
                self.stdout.write(self.style.WARNING(
                    f'⚠ Klaidų: {error_count}'
                ))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Klaida apdorojant failą: {str(e)}'))

