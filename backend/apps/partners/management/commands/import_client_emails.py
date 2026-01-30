import os
import re
from django.core.management.base import BaseCommand
from apps.partners.models import Partner, Contact
from difflib import SequenceMatcher
import openpyxl


class Command(BaseCommand):
    help = 'Importuoja email adresus iš Excel failo ir prideda juos prie klientų'

    def add_arguments(self, parser):
        parser.add_argument(
            'file',
            type=str,
            help='Kelias į Excel failą (infotrans_uzsakymai.xlsx)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Tik parodyti, ką būtų padaryta, bet nieko nekeisti',
        )
        parser.add_argument(
            '--min-similarity',
            type=float,
            default=0.85,
            help='Minimalus panašumas kliento pavadinimams (0.0-1.0, default: 0.85)',
        )

    def similarity(self, a, b):
        """Apskaičiuoja panašumą tarp dviejų eilučių"""
        if not a or not b:
            return 0
        return SequenceMatcher(None, str(a).lower().strip(), str(b).lower().strip()).ratio()

    def clean_company_name(self, name):
        """Išvalo įmonės pavadinimą (pašalina UAB, SIA, MB, ir t.t.)"""
        if not name:
            return ''
        cleaned = re.sub(r'\b(UAB|SIA|MB|AB|VšĮ|IĮ)\b\.?', '', str(name), flags=re.IGNORECASE)
        return cleaned.strip(' ,.')

    def extract_emails(self, text):
        """Ištraukia email adresus iš teksto (net jei tai tik "info@" arba "invoices@")"""
        if not text:
            return []
        
        emails = []
        text_str = str(text)
        
        # Pirmiausia ištraukti pilnus email adresus (pvz., info@aksola.lt)
        full_email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        full_emails = re.findall(full_email_pattern, text_str)
        emails.extend(full_emails)
        
        # Tada pašalinti pilnus email adresus iš teksto, kad nerastume jų dar kartą
        text_without_full = text_str
        for full_email in full_emails:
            text_without_full = text_without_full.replace(full_email, ' ')
        
        # Dabar ieškome dalinių email adresų (pvz., "info@" arba "invoices@")
        # Ieškome teksto, kuris baigiasi su @ (bet nėra pilnas email)
        partial_email_pattern = r'[a-zA-Z0-9._%+-]+@(?![a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
        partial_matches = re.findall(partial_email_pattern, text_without_full)
        for partial in partial_matches:
            cleaned = partial.strip()
            if cleaned and cleaned not in emails:
                emails.append(cleaned)
        
        # Pašalinti dublikatus, bet išlaikyti tvarką
        seen = set()
        unique_emails = []
        for email in emails:
            email_lower = email.lower()
            if email_lower not in seen:
                seen.add(email_lower)
                unique_emails.append(email)
        
        return unique_emails

    def find_client_by_name(self, client_name, min_similarity=0.85):
        """Randa klientą duomenų bazėje pagal pavadinimą (fuzzy matching)"""
        if not client_name:
            return None
        
        all_clients = Partner.objects.filter(is_client=True)
        clean_name = self.clean_company_name(client_name)
        
        best_match = None
        best_score = 0
        
        for client in all_clients:
            clean_client_name = self.clean_company_name(client.name)
            similarity_score = self.similarity(clean_name, clean_client_name)
            
            if similarity_score > best_score:
                best_score = similarity_score
                best_match = client
        
        if best_score >= min_similarity:
            return best_match, best_score
        
        return None, 0

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options.get('dry_run', False)
        min_similarity = options.get('min_similarity', 0.85)
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'Failas nerastas: {file_path}'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'Nuskaitomas failas: {file_path}'))
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Klaida atidarant failą: {e}'))
            return
        
        stats = {
            'total_rows': 0,
            'matched_clients': 0,
            'emails_added': 0,
            'emails_skipped': 0,
            'clients_not_found': 0,
            'errors': 0
        }
        
        self.stdout.write(f'\nPradedamas importavimas...')
        self.stdout.write(f'Iš viso eilučių: {ws.max_row - 1}')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN MODE - nieko nebus pakeista!\n'))
        
        # Eiti per visas eilutes (praleisti header)
        total_rows = ws.max_row - 1
        for row_num in range(2, ws.max_row + 1):
            # Rodyti progress kas 100 eilučių
            if (row_num - 2) % 100 == 0:
                progress = ((row_num - 2) / total_rows) * 100
                self.stdout.write(f'Apdorota: {row_num - 2} / {total_rows} eilučių ({progress:.1f}%)...')
            client_name = ws.cell(row=row_num, column=2).value  # Stulpelis B
            email_text = ws.cell(row=row_num, column=25).value  # Stulpelis Y (25)
            
            if not client_name or not email_text:
                continue
            
            stats['total_rows'] += 1
            client_name_str = str(client_name).strip()
            
            # Rasti klientą
            client, similarity_score = self.find_client_by_name(client_name_str, min_similarity)
            
            if not client:
                stats['clients_not_found'] += 1
                if stats['clients_not_found'] <= 5:  # Rodyti tik pirmus 5
                    self.stdout.write(
                        self.style.WARNING(f'Eilutė {row_num}: Klientas nerastas: {client_name_str}')
                    )
                continue
            
            stats['matched_clients'] += 1
            
            # Ištraukti email adresus
            emails = self.extract_emails(email_text)
            
            if not emails:
                continue
            
            # Pridėti email adresus prie kliento
            for email in emails:
                email_lower = email.lower().strip()
                
                # Patikrinti, ar email jau egzistuoja
                existing_contact = Contact.objects.filter(
                    partner=client,
                    email__iexact=email_lower
                ).first()
                
                if existing_contact:
                    stats['emails_skipped'] += 1
                    continue
                
                # Patikrinti, ar email jau yra kaip contact_person
                if client.contact_person and client.contact_person.email:
                    if client.contact_person.email.lower().strip() == email_lower:
                        stats['emails_skipped'] += 1
                        continue
                
                # Sukurti naują Contact objektą
                # Nesvarbu, ar email validus ar ne - vis tiek įrašome į email lauką
                if not dry_run:
                    try:
                        contact = Contact.objects.create(
                            partner=client,
                            email=email_lower,  # Įrašome kaip yra, net jei nevalidus
                            first_name='',  # Tuščias, nes nežinome vardo
                            last_name='',   # Tuščias, nes nežinome pavardės
                            notes=f'Importuotas iš Excel failo, eilutė {row_num}'
                        )
                        stats['emails_added'] += 1
                        self.stdout.write(
                            self.style.SUCCESS(f'Eilutė {row_num}: Pridėtas email {email_lower} prie {client.name}')
                        )
                    except Exception as e:
                        stats['errors'] += 1
                        self.stdout.write(
                            self.style.ERROR(f'Eilutė {row_num}: Klaida kuriant kontaktą: {e}')
                        )
                else:
                    stats['emails_added'] += 1
                    self.stdout.write(
                        f'  [DRY RUN] Būtų pridėtas email {email_lower} prie {client.name}'
                    )
        
        # Rodyti statistiką
        self.stdout.write(self.style.SUCCESS(f'\n=== IMPORT AVIMO STATISTIKA ==='))
        self.stdout.write(f'Iš viso eilučių apdorota: {stats["total_rows"]}')
        self.stdout.write(f'Rastų klientų: {stats["matched_clients"]} ({stats["matched_clients"]/stats["total_rows"]*100:.1f}%)')
        self.stdout.write(f'Pridėta email adresų: {stats["emails_added"]}')
        self.stdout.write(f'Praleista email adresų (jau egzistuoja): {stats["emails_skipped"]}')
        self.stdout.write(f'Klientų nerasta: {stats["clients_not_found"]}')
        if stats['errors'] > 0:
            self.stdout.write(self.style.ERROR(f'Klaidų: {stats["errors"]}'))
        
        if dry_run:
            self.stdout.write(self.style.WARNING('\nDRY RUN MODE - nieko nebuvo pakeista!'))
            self.stdout.write(self.style.SUCCESS('Paleiskite be --dry-run, kad išsaugotumėte pakeitimus.'))
        else:
            self.stdout.write(self.style.SUCCESS('\nImportavimas baigtas!'))

