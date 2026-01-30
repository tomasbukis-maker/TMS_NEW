"""
Partnerių importavimo pagalbinės funkcijos
"""
import csv
from openpyxl import load_workbook
from .models import Partner, Contact


def safe_get(row, key, default=''):
    """Saugus duomenų gavimas iš dict"""
    value = row.get(key, default)
    return str(value).strip() if value is not None else default


def parse_contact_name(kontaktiniai_asmenys):
    """Parsina kontaktinio asmens vardą ir pavardę"""
    if not kontaktiniai_asmenys:
        return 'Nenurodytas', ''
    
    name_parts = kontaktiniai_asmenys.replace(';', ',').split(',')
    first_name = name_parts[0].strip() if len(name_parts) > 0 else ''
    last_name = name_parts[1].strip() if len(name_parts) > 1 else (kontaktiniai_asmenys.strip() if not first_name else '')
    
    if not first_name and not last_name:
        first_name = 'Nenurodytas'
    
    return first_name, last_name


def import_partners_from_file(file, is_client=True, is_supplier=False, update_existing=False):
    """
    Importuoja partnerius iš XLSX arba CSV failo.
    
    Returns:
        dict su 'imported', 'skipped', 'errors', 'results' laukais
    """
    imported_count = 0
    skipped_count = 0
    error_count = 0
    results = []
    
    file_name = file.name.lower()
    is_xlsx = file_name.endswith('.xlsx')
    
    try:
        if is_xlsx:
            # Apdorojame XLSX
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active
            
            # Gauname header eilutę
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            # Konvertuojame į dict formatą
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_dict = {}
                for idx, cell in enumerate(row):
                    header_name = headers[idx] if idx < len(headers) else ''
                    row_dict[header_name] = str(cell.value) if cell.value else ''
                rows.append(row_dict)
        else:
            # Apdorojame CSV
            file.seek(0)
            # Bandoma nustatyti delimiter iš pirmų eilučių
            sample_lines = []
            for i, line in enumerate(file):
                if isinstance(line, bytes):
                    line = line.decode('utf-8')
                sample_lines.append(line)
                if i >= 5:  # Pirmos 5 eilutės
                    break
            
            file.seek(0)
            sample = ''.join(sample_lines)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter
            
            # Nuskaitome visą failą
            lines = []
            for line in file:
                if isinstance(line, bytes):
                    line = line.decode('utf-8')
                lines.append(line)
            
            reader = csv.DictReader(lines, delimiter=delimiter)
            rows = list(reader)
        
        # Apdorojame eilutes
        for row_num, row in enumerate(rows, start=2):
            try:
                firmos_pavadinimas = safe_get(row, 'Firmos pavadinimas')
                imones_kodas = safe_get(row, 'Imones kodas')
                
                # Jei nėra firmos pavadinimo, naudojame default
                if not firmos_pavadinimas:
                    firmos_pavadinimas = f'Partneris #{row_num}'  # Laikinas pavadinimas
                
                # Jei nėra įmonės kodo, generuojame unikalų
                if not imones_kodas:
                    # Generuojame unikalų kodą pagal pavadinimą arba eilutės numerį
                    base_code = firmos_pavadinimas.upper().replace(' ', '')[:10] if firmos_pavadinimas else f'PART{row_num}'
                    counter = 1
                    imones_kodas = f'{base_code}{counter:03d}'
                    
                    # Tikriname ar toks kodas jau egzistuoja
                    while Partner.objects.filter(code=imones_kodas).exists():
                        counter += 1
                        imones_kodas = f'{base_code}{counter:03d}'
                    
                    results.append({
                        'row': row_num,
                        'status': 'imported',
                        'warning': f'Trūksta įmonės kodo - sugeneruotas: {imones_kodas}'
                    })
                
                # Tikriname ar egzistuoja tik jei yra kodas
                existing = None
                if imones_kodas:
                    existing = Partner.objects.filter(code=imones_kodas).first()
                
                if existing:
                    if not update_existing:
                        results.append({
                            'row': row_num,
                            'status': 'skipped',
                            'message': f'Partneris su kodu {imones_kodas} jau egzistuoja (naudokite "Atnaujinti egzistuojančius")'
                        })
                        skipped_count += 1
                        continue
                    else:
                        partner = existing
                        created = False
                else:
                    partner = None
                    created = True
                
                # Formuojame adresą
                salis = safe_get(row, 'Salis')
                miestas = safe_get(row, 'Miestas')
                address_parts = [miestas, salis]
                address = ', '.join([p for p in address_parts if p])
                
                # Kontaktinis asmuo
                contact_person = None
                kontaktiniai_asmenys = safe_get(row, 'Kontaktiniai asmenys')
                el_pastas = safe_get(row, 'El. pastas')
                telefonas = safe_get(row, 'Telefonas')
                
                if kontaktiniai_asmenys or el_pastas or telefonas:
                    first_name, last_name = parse_contact_name(kontaktiniai_asmenys)
                    
                    contact_person, _ = Contact.objects.get_or_create(
                        email=el_pastas,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'phone': telefonas if telefonas else '',
                        }
                    )
                
                # Sukuriame arba atnaujiname partnerį
                partner_data = {
                    'name': firmos_pavadinimas,
                    'vat_code': safe_get(row, 'PVM kodas'),
                    'address': address,
                    'contact_person': contact_person,
                    'is_client': is_client,
                    'is_supplier': is_supplier,
                    'status': 'active',
                    'payment_term_days': 30,
                }
                
                if created:
                    partner = Partner.objects.create(code=imones_kodas, **partner_data)
                    imported_count += 1
                    results.append({
                        'row': row_num,
                        'status': 'imported',
                        'partner_id': partner.id,
                        'message': f'Importuotas: {firmos_pavadinimas} ({imones_kodas})'
                    })
                else:
                    for key, value in partner_data.items():
                        setattr(partner, key, value)
                    partner.code = imones_kodas  # Code negali keisti
                    partner.save()
                    imported_count += 1
                    results.append({
                        'row': row_num,
                        'status': 'updated',
                        'partner_id': partner.id,
                        'message': f'Atnaujintas: {firmos_pavadinimas} ({imones_kodas})'
                    })
                    
            except Exception as e:
                error_count += 1
                results.append({
                    'row': row_num,
                    'status': 'error',
                    'message': f'Klaida: {str(e)}'
                })
        
        return {
            'success': True,
            'imported': imported_count,
            'skipped': skipped_count,
            'errors': error_count,
            'results': results[:100],  # Tik pirmi 100 rezultatų
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Klaida apdorojant failą: {str(e)}',
            'imported': imported_count,
            'skipped': skipped_count,
            'errors': error_count,
            'results': results,
        }

